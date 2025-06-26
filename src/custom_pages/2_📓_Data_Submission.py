import streamlit as st
import pandas as pd
from typing import List, Optional, Dict
from influxdb_client_3 import InfluxDBClient3, flight_client_options
import logging
import certifi
import seaborn as sns
import plotly.express as px
from config import INIT_PARAMS as IPS
from config.loader import load_config
import os
os.environ["STREAMLIT_SERVER_RUN_ON_SAVE"] = "false"
import certifi
import matplotlib.pyplot as plt
from datetime import datetime, timedelta, timezone, time
import pytz
from io import BytesIO
local_tz = pytz.timezone("Asia/Kolkata")  # or use your actual timezone

from dotenv import load_dotenv
load_dotenv() 

# Our custom utilities
from utils.data_submission import (
    layout_sub_sections,
    strip_section_prefix,
    flatten_single_sample_df,
    flatten_multi_sample_df_agg,
    create_production_report_xlsx
)




from dotenv import load_dotenv
load_dotenv() 

# Our custom utilities
from utils.data_submission import (
    layout_sub_sections,
    strip_section_prefix,
    flatten_single_sample_df,
    flatten_multi_sample_df_agg,
    create_production_report_xlsx
)


# --------------------------------------------------------------------------------
# HELPER FUNCTION: Submitting data for a single date
# --------------------------------------------------------------------------------
def submit_data_for_date(
    selected_date: pd.Timestamp,
    all_section_data: dict[str, dict[str, pd.DataFrame]],
    existing_df: pd.DataFrame,
):
    """
    Merges new data for the chosen date into existing_df (st.session_state["submitted_data"]).
    If the date already exists, update only the columns that have new (non-NaN) numeric values.
    If it's a new date, append a new row.

    :param selected_date: The date chosen by the user
    :param all_section_data:
        Dict of { section_key: { sub_section_name: DataFrame } }
        from layout_sub_sections. Each sub-section is either single or multi-sample.
    :param existing_df: The current DataFrame in st.session_state["submitted_data"].
    :return:
        (updated_df, changed_cols)
        updated_df: The updated DataFrame with the new columns/values
        changed_cols: A list of columns that were updated or newly created
    """
    row_list = []
    # Flatten or aggregate each sub-section
    for sec_key, sub_dict in all_section_data.items():
        for sub_name, df in sub_dict.items():
            # Remove "a. ", "b. ", etc. from sub-section name
            clean_name = strip_section_prefix(sub_name)

            # If it's multi-sample => flatten with avg/min/max
            if df.shape[0] > 1:
                flattened = flatten_multi_sample_df_agg(df, prefix=clean_name)
            else:
                flattened = flatten_single_sample_df(df, prefix=clean_name)

            row_list.append(flattened)

    # If no data, return without changes
    if not row_list:
        return existing_df, []

    # Combine horizontally => one row (index = selected_date)
    combined_row = pd.concat(row_list, axis=1)
    combined_row.index = [selected_date]

    changed_cols = []

    # Check if this date is already in the existing DataFrame
    if selected_date in existing_df.index:
        # We are updating an existing row
        old_row = existing_df.loc[[selected_date]].copy()

        # Merge new values in. If the new value is numeric (and not empty),
        # overwrite old value if old was NaN or different.
        for col in combined_row.columns:
            new_val = combined_row.loc[selected_date, col]
            old_val = old_row.loc[selected_date, col] if col in old_row.columns else None

            # If new_val is not NaN or empty, we update
            if pd.notna(new_val) and str(new_val).strip() != "":
                if pd.isna(old_val) or old_val != new_val:
                    old_row.loc[selected_date, col] = new_val
                    changed_cols.append(col)
            else:
                # If new_val is empty or NaN, do not overwrite
                pass

        # Put updated row back into existing_df
        for col in old_row.columns:
            existing_df.loc[selected_date, col] = old_row.loc[selected_date, col]

    else:
        # Brand-new date => just append
        existing_df = pd.concat([existing_df, combined_row], axis=0)
        changed_cols = list(combined_row.columns)

    return existing_df, changed_cols

def download_production_report_button(data_dict, year):
    """
    Creates an in-memory Excel file for the production report
    and offers a download button in Streamlit.
    """
    excel_data = create_production_report_xlsx(data_dict, year)
    st.download_button(
        label="Download Production Report",
        data=excel_data,
        file_name="production_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --------------------------------------------------------------------------------
# MAIN PAGE CODE
# --------------------------------------------------------------------------------
config = load_config("setting_submission.yaml")
log = logging.getLogger("root")

# Ensure session_state data exists
if "submitted_data" not in st.session_state:
    st.session_state["submitted_data"] = pd.DataFrame()

st.title("Offline Data Submission Portal")

# Select date
selected_date = st.date_input("Select Date for Submission", value=pd.Timestamp.now())
st.markdown("---")

# Build Tabs from config
section_names = list(config["sections"].keys())
tabs = st.tabs([config["sections"][sec]["display_title"] for sec in section_names])

all_section_data = {}
for idx, sec_key in enumerate(section_names):
    with tabs[idx]:
        section_cfg = config["sections"][sec_key]
        st.write(f"**Section:** {section_cfg['display_title']}")
        sub_sections = section_cfg["sub_sections"]

        # Layout the sub-sections for user input
        sub_data_dict = layout_sub_sections(sub_sections, selected_date)
        all_section_data[sec_key] = sub_data_dict

st.markdown("---")
st.subheader("Data submission")
# --------------------------------------------------------------------------------
# SUBMIT LOGIC - Merging new data or creating a new row
# --------------------------------------------------------------------------------
if st.button("Submit", help="Submits data across all tabs"):
    updated_df, changed_cols = submit_data_for_date(
        selected_date=selected_date,
        all_section_data=all_section_data,
        existing_df=st.session_state["submitted_data"]
    )
    st.session_state["submitted_data"] = updated_df

    if changed_cols:
        st.success(f"Data for submitted/updated!")
        st.write("### Columns Changed or Filled:")
        st.write(changed_cols)
    else:
        st.warning("No changes were made. (Either all fields were empty or identical.)")

    st.write("### All Data After Submission:")
    st.dataframe(st.session_state["submitted_data"])

# --------------------------------------------------------------------------------
# VIEW ALL DATA
# --------------------------------------------------------------------------------
if st.button("View All Submitted Data"):
    st.write("### All Submitted Data:")
    if st.session_state["submitted_data"].empty:
        st.info("No data has been submitted yet.")
    else:
        st.dataframe(st.session_state["submitted_data"])

st.markdown("---")

# --------------------------------------------------------------------------------
# DOWNLOAD EXCEL DEMO (Production & placeholders for other categories)
# --------------------------------------------------------------------------------
st.subheader("Download data as excel files")
cols = st.columns(3)
with cols[0]:
    # This is just a stub example showing how you'd create a data_dict 
    # (month->section->parameters->(date->value)) from st.session_state["submitted_data"].
    # Right now st.session_state["submitted_data"] is a wide table with columns = param_names.
    # You must convert it to a monthly dictionary if you want the create_production_report_xlsx style.
    # For now, just show a dummy data_dict example, or adapt your code to build it from the DataFrame.
    dummy_data_dict = {
        "May": {
            "Production Product": {
                "Total Hot Metal Production (MT)": {
                    "2023-05-01": 123.0,
                    "2023-05-02": 140.5
                },
                "Slag Generation (Calculated) (MT)": {
                    "2023-05-01": 90.0,
                    "2023-05-02": 95.5
                }
            },
            "By-Product Generation": {
                "BF(RF) Coke Breeze -16mm (MT)": {
                    "2023-05-01": 12,
                    "2023-05-02": 13
                }
            }
        }
    }

    if st.button("Production Report", help="Download the Daily Production Report, custom template."):
        download_production_report_button(dummy_data_dict, 2023)

with cols[1]:
    if st.button("Raw Material Analysis", help="Download Raw material analysis, custom template."):
        st.write("Placeholder for raw material analysis download.")
        # You can adapt a function like `create_production_report_xlsx`
        # but with a layout for raw material analysis, then st.download_button

with cols[2]:
    if st.button("Slag/Hot Metal Chemistry", help="Download Slag & Hot Metal Chemistry"):
        st.write("Placeholder for Slag/HM Chemistry download.")
        # Similarly adapt or implement another writer function for slag/hot metal


st.title("Visualisation tool")

with st.expander("How to Use DataWalker", expanded=False):
    st.markdown("""
    **DataWalker** is a tool for interactive exploration of tabular data. You can:
    - Use the **global search** bar to find any specific term across the dataset.
    - Apply filters, sort columns, and navigate through data interactively.
    - Enable or disable columns for streamlined viewing.
    """)

#--------------------------------------------------------------------------
st.subheader("Distribution plots")
cols = st.columns(2)
df = pd.read_excel("src/data/Cleaned_Dataset_BF_hourly (2).xlsx")
with cols[0]:
    x = st.selectbox("Select X feature", df.columns[1:])

with cols[1]:
    y = st.selectbox("Select Y feature", IPS.OUTPUT_PARAMS)
fig = sns.jointplot(data=df, x=x, y=y, kind='reg')
st.pyplot(fig, use_container_width=False)


#--------------------------------------------------------------------------
st.subheader("Timeseries plot")
cols = st.columns(2)
df = pd.read_excel("src/data/Cleaned_Dataset_BF_hourly (2).xlsx")
y = []
cols_y = df.columns[1:]
with cols[0]:
    y.append(st.selectbox("Select feature 1", cols_y))
    
with cols[1]:
    remain_cols = cols_y.insert(0, 'None')
    y.append(st.selectbox("Select feature 2", remain_cols))

y = [i for i in y if i != 'None']
# Plot the data
fig = px.line(
    df,
    x=df.columns[0],
    y=y,
    hover_data={df.columns[0]: "|%B %d, %Y"},
    markers=True
)

# Display the plot
st.plotly_chart(fig, use_container_width=True)

# --------------------------------------------------------------------------------
# SHOW 6 MEASUREMENTS FROM INFLUXDB
# --------------------------------------------------------------------------------

# load once
config = load_config()

FIELD_LABELS = {
    internal_key: human_label
    for mapping in config["data_mapping"].values()
    for human_label, internal_key in mapping.items()
}

# InfluxDB connection settings

influx_config = config.get("influxdb", {})
host = influx_config.get("host", "")
org = influx_config.get("org", "")
database = influx_config.get("database", "")
token = os.getenv("INFLUX_TOKEN", "")

with open(certifi.where(), "r") as f:
    cert = f.read()


# Parse TIME_OPTIONS from config
TIME_OPTIONS = {
    label: timedelta(seconds=secs)
    for label, secs in config.get("time_options", {}).items()
}

# Parse FREQUENCY_OPTIONS from config
FREQUENCY_OPTIONS = {
    label: (info["label"], timedelta(seconds=info["duration"]))
    for label, info in config.get("frequency_options", {}).items()
}

MEASUREMENT_LABELS = config.get("measurement_labels", {})



def get_measurements():
    client = InfluxDBClient3(host=host, org=org, token=token, database=database,
                             flight_client_options=flight_client_options(tls_root_certs=cert))
    df = client.query("SHOW TABLES", mode="pandas")
    client.close()
    return df[df["table_type"] == "BASE TABLE"]["table_name"].tolist()

def get_fields(measurement):
    query = f"SELECT * FROM {measurement} LIMIT 1"
    client = InfluxDBClient3(host=host, org=org, token=token, database=database,
                             flight_client_options=flight_client_options(tls_root_certs=cert))
    df = client.query(query, mode="pandas")
    client.close()
    return [col for col in df.columns if col != "time"]

def get_data(measurement, fields, time_range):
    end = datetime.now(timezone.utc)
    start = end - TIME_OPTIONS[time_range]
    start_iso = start.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
    end_iso = end.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
    field_str = ", ".join(fields)

    query = f"SELECT time, {field_str} FROM {measurement} WHERE time >= timestamp '{start_iso}' AND time <= timestamp '{end_iso}'"
    client = InfluxDBClient3(host=host, org=org, token=token, database=database,
                             flight_client_options=flight_client_options(tls_root_certs=cert))
    df = client.query(query=query, mode="pandas")
    client.close()

    if not df.empty:
        df["time"] = pd.to_datetime(df["time"]).dt.tz_localize('UTC').dt.tz_convert('Asia/Kolkata').dt.tz_localize(None)
    return df


def average_data(df, freq, time_delta):
    df["time"] = pd.to_datetime(df["time"])
    df = df.sort_values("time")

    # Strictly use selected time range
    end_time = df["time"].max()
    start_time = end_time - time_delta

    # Filter only the selected time window
    df = df[df["time"] > start_time]

    # Resample using exact start time as origin
    df.set_index("time", inplace=True)
    df_avg = df.resample(freq, origin=start_time, label='right', closed='right').mean().dropna().reset_index()
    
    return df_avg







# --- Streamlit UI ---
st.title("📊 BF2 Data Downloader")

if "measurements" not in st.session_state:
    st.session_state.measurements = []
if "selected_measurements" not in st.session_state:
    st.session_state.selected_measurements = set()
if "avg_mode" not in st.session_state:
    st.session_state.avg_mode = False
if "select_all" not in st.session_state:
    st.session_state.select_all = False

if st.button("📂 Show Measurements", key="show_measurements_button"):
    st.session_state.measurements = get_measurements()

if st.session_state.measurements:
    st.subheader("Select Measurements")

    col1, col2 = st.columns(2)
    

    with col1:
        st.toggle("Select All", key="select_all")
        if st.session_state.select_all:
            st.session_state.selected_measurements = set(st.session_state.measurements)
        else:
            st.session_state.selected_measurements = set()

    with col2:
        st.toggle("Average", key="avg_mode")

    cols = st.columns(4)
    for i, meas in enumerate(st.session_state.measurements):
        col = cols[i % 4]  # Rotate across 4 columns
        label = MEASUREMENT_LABELS.get(meas, meas)
        with col:
            if st.checkbox(label, value=meas in st.session_state.selected_measurements, key=meas):
                st.session_state.selected_measurements.add(meas)
            else:
                st.session_state.selected_measurements.discard(meas)


if st.session_state.selected_measurements:
    time_range = st.selectbox("Select Time Range:", list(TIME_OPTIONS.keys()))

    # 1) pick averaging frequency (if any)
    freq = None
    if st.session_state.avg_mode:
        freq_label = st.selectbox("⏱️ Average Frequency:", list(FREQUENCY_OPTIONS.keys()))
        freq_str, freq_delta = FREQUENCY_OPTIONS[freq_label]
        if freq_delta > TIME_OPTIONS[time_range]:
            st.error("❌ Frequency is greater than time range.")
            freq = None
        else:
            freq = freq_str

    # 2) on button click: fetch, rename & merge in one pass
    if st.button("⬇️ Show Excel File", key="download_combined_csv"):
        combined_df = pd.DataFrame()

        for meas in st.session_state.selected_measurements:
            # fetch raw (or averaged) data
            df = get_data(meas, get_fields(meas), time_range)
            if df.empty:
                continue
            if freq:
                df = average_data(df, freq, TIME_OPTIONS[time_range])

            # rename cols using FIELD_LABELS + MEASUREMENT_LABELS
            df = df.rename(columns={
                col: f"{MEASUREMENT_LABELS.get(meas, meas)} - {FIELD_LABELS.get(col, col)}"
                for col in df.columns if col != "time"
            })

            # merge into master
            combined_df = df if combined_df.empty else combined_df.merge(df, on="time", how="outer")

        # 3) build filename
        prefix = "avg_" if (st.session_state.avg_mode and freq) else "raw_"
        all_sel = st.session_state.selected_measurements
        sheet_part = (
            "combined"
            if len(all_sel) == len(st.session_state.measurements)
            else "_".join(sorted(all_sel)).replace(" ", "_")
        )
        file_name = f"{prefix}data_{sheet_part}_{time_range.replace(' ', '_')}.csv"

        # 4) show & download
        if not combined_df.empty:
            combined_df = combined_df.sort_values("time").reset_index(drop=True)
            combined_df.index += 1
            st.dataframe(combined_df)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            combined_df.to_excel(writer, index=True, index_label="Index", sheet_name="BF2 Data")
            
        excel_data = output.getvalue()

        st.download_button(
            "📄 Download Excel File",
            data=excel_data,
            file_name=file_name.replace(".csv", ".xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )




# --------------------------------------------------------------------------------
# Hourly avg for process parameters
# ------------------------------------------------------------------------



field_mapping = config.get("field_mapping", {})
# InfluxDB settings
influx_config = config.get("influxdb", {})
host = influx_config.get("host", "")
org = influx_config.get("org", "")
database = influx_config.get("database", "")

token = os.getenv("INFLUX_TOKEN", "")

with open(certifi.where(), "r") as f:
    cert = f.read()


REQUIRED_FIELDS = [v for v in field_mapping.values() if v]



# --- Functions ---
def get_measurements():
    client = InfluxDBClient3(host=host, org=org, token=token, database=database,
                             flight_client_options=flight_client_options(tls_root_certs=cert))
    df = client.query("SHOW TABLES", mode="pandas")
    client.close()
    return df[df["table_type"] == "BASE TABLE"]["table_name"].tolist()

def get_fields(measurement):
    query = f"SELECT * FROM {measurement} LIMIT 1"
    client = InfluxDBClient3(host=host, org=org, token=token, database=database,
                             flight_client_options=flight_client_options(tls_root_certs=cert))
    df = client.query(query, mode="pandas")
    client.close()
    return [col for col in df.columns if col != "time"]

def get_data(measurement, fields, start, end):
    start_iso = start.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
    end_iso = end.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
    field_str = ", ".join(fields)

    query = f"""
        SELECT time, {field_str}
        FROM {measurement}
        WHERE time >= timestamp '{start_iso}' AND time <= timestamp '{end_iso}'
    """
    client = InfluxDBClient3(host=host, org=org, token=token, database=database,
                             flight_client_options=flight_client_options(tls_root_certs=cert))
    df = client.query(query=query, mode="pandas")
    client.close()
    return df


def average_data(df):
    # 1️⃣ Parse as UTC
    df["time"] = pd.to_datetime(df["time"], utc=True)
    
    # 2️⃣ Make it the index and convert to India time
    df = df.set_index("time").tz_convert("Asia/Kolkata")
    
    # 3️⃣ Resample on the India‐time index
    df_hourly = (
        df
        .resample("1h", label="right", closed="right")
        .mean()
        .dropna()
    )
    
    # 4️⃣ (Optional) Remove the timezone info so Excel shows plain times
    df_hourly.index = df_hourly.index.tz_localize(None)
    
    # 5️⃣ Reset index so “time” comes back as a column
    return df_hourly.reset_index()


def make_utc_bounds_for_ist_day(selected_date):
    """
    Given a date (naive), returns (start_utc, end_utc) so that
    in IST it covers 00:00→23:59:59 on that date.
    """
    india = pytz.timezone("Asia/Kolkata")

    # 1) Localize to IST midnight and IST 23:59:59
    start_ist = india.localize(datetime.combine(selected_date, time(0, 0)))
    end_ist   = india.localize(datetime.combine(selected_date, time(23, 59, 59)))

    # 2) Convert those instants to UTC (what InfluxDB needs)
    start_utc = start_ist.astimezone(pytz.UTC)
    end_utc   = end_ist.astimezone(pytz.UTC)

    return start_utc, end_utc




# --- Streamlit UI ---
st.title("📊 Hourly Average Data for Process Parameters")

# Unified single-date input (defaults to today)
selected_date = st.date_input("📅 Select a date to download full-day hourly data:", value=datetime.utcnow().date())

# Compute InfluxDB query bounds so that in IST we cover 00:00→23:59:59
start_time, end_time = make_utc_bounds_for_ist_day(selected_date)

# Show time range (UTC)


# Fetch and download section
if st.button("🔄 Fetch & Download Hourly Averages for the Day"):
    combined_df = pd.DataFrame()
    measurements = get_measurements()

    for meas in measurements:
        fields = get_fields(meas)
        matched_fields = [f for f in fields if f in REQUIRED_FIELDS]

        if not matched_fields:
            continue

        df = get_data(meas, matched_fields, start_time, end_time)
        if df.empty:
            continue

        df = average_data(df)
        df.columns = ['time'] + [col for col in df.columns if col != 'time']
        combined_df = df if combined_df.empty else combined_df.merge(df, on="time", how="outer")

    if not combined_df.empty:
        combined_df = combined_df.sort_values("time").reset_index(drop=True)
        field_order = list(field_mapping.keys())
        output_df = pd.DataFrame()
        output_df["Time,"] = combined_df["time"] if "time" in combined_df else pd.NaT

        for label in field_order:
            internal = field_mapping[label]
            if internal:
                # Get data from combined_df if present
                if internal in combined_df.columns:
                    output_df[label] = combined_df[internal].fillna("")
                else:
                    output_df[label] = ""
            else:
                # Leave column untouched so Excel formula remains
                output_df[label] = ["" for _ in range(len(output_df))]


        output_df.index += 1
        st.dataframe(output_df)

        # Load Excel template and populate data
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        template_path = "src/templates/examplebf2process download.xlsx"
        output_df["Time,"] = pd.to_datetime(output_df["Time,"]).dt.strftime("%H:%M")
        wb = load_workbook(template_path)
        ws = wb.active
        # 🗓️ Write year in B2 cell
        ws["B2"] = selected_date.strftime("%d-%b-%Y")  # e.g., 2025-06-24
        start_row, start_col = 7, 2
        for r_idx, row in enumerate(dataframe_to_rows(output_df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                # Get column name from output_df
                col_name = output_df.columns[c_idx - start_col]
                internal = field_mapping.get(col_name, "")

                # Only write value if the column is mapped and value is valid
                if internal and pd.notna(value):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                # Else skip writing — this preserves existing formulas in the cell
                

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        start_str = start_time.strftime("%Y-%m-%d_%H-%M")
        end_str = end_time.strftime("%Y-%m-%d_%H-%M")
        filename = f"bf2processparamsofhourlyavg_{start_str}_to_{end_str}.xlsx"

        st.download_button(
            label="⬇️ Download Excel File",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("⚠️ No data available for the selected date.")
