import streamlit as st
import pandas as pd
import numpy as np
import re

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

def parse_parameter_and_unit(full_param: str):
    """
    Splits a parameter string like 'Total Hot Metal Production (MT)'
    into:
       param_name = 'Total Hot Metal Production'
       unit       = 'MT'
    If no bracket is found, returns (full_param, "")
    """
    match = re.search(r'\(([^)]*)\)\s*$', full_param)
    if match:
        unit = match.group(1)  # text inside ( )
        param_name = full_param[: match.start()].strip()
        return param_name, unit
    else:
        return full_param.strip(), ""

def create_production_report_xlsx(data_dict, year):
    """
    Create an Excel file (as bytes) with one worksheet per month.
    The 'data_dict' is assumed to have this structure:
        {
          "May": {
             "Production Product": {
                "Total Hot Metal Production (MT)": {
                    "2023-05-01": 123.0,
                    "2023-05-02": 140.5,
                    ...
                },
                "Slag Generation (Calculated) (MT)": {...},
                ...
             },
             "By-Product Generation": {...},
             ...
          },
          "June": {
             ...
          }
        }
    'year': e.g. 2023

    Returns: BytesIO representing the Excel file
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for month_name, month_data in data_dict.items():
        # month_data => { section_name: { param_name: { date_str: val, ... }, ... }, ... }
        sheet_name = f"{month_name}-{str(year)[-2:]}"
        ws = wb.create_sheet(title=sheet_name)

        # Optional styling for columns
        ws.column_dimensions["B"].width = 20  # Section name
        ws.column_dimensions["C"].width = 40  # Parameter name
        ws.column_dimensions["D"].width = 10  # Units

        # Gather all dates used in this month's data
        all_dates = set()
        for section_name, section_params in month_data.items():
            for param_name, date_dict in section_params.items():
                for date_str in date_dict.keys():
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                    all_dates.add(dt)

        all_dates = sorted(list(all_dates))

        # Write date headers in row 3, columns E onward
        start_col_for_dates = 5
        for i, dt in enumerate(all_dates):
            col_index = start_col_for_dates + i
            col_letter = get_column_letter(col_index)
            cell = ws.cell(row=3, column=col_index, value=dt.strftime("%d/%m"))
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[col_letter].width = 10

        current_row = 4

        # Fill data section by section
        for section_name, section_params in month_data.items():
            section_start_row = current_row

            for param_name, date_dict in section_params.items():
                # parse out param vs. unit
                param_str, unit_str = parse_parameter_and_unit(param_name)

                ws.cell(row=current_row, column=3, value=param_str)
                ws.cell(row=current_row, column=4, value=unit_str)

                # fill date values
                for i, dt in enumerate(all_dates):
                    col_index = start_col_for_dates + i
                    val = date_dict.get(dt.strftime("%Y-%m-%d"), None)
                    if val is not None:
                        ws.cell(row=current_row, column=col_index, value=val)

                current_row += 1

            end_row = current_row - 1
            if section_start_row <= end_row:
                ws.merge_cells(start_row=section_start_row, start_column=2,
                               end_row=end_row, end_column=2)
                ws.cell(row=section_start_row, column=2, value=section_name)
                ws.cell(row=section_start_row, column=2).alignment = Alignment(
                    vertical="center", horizontal="center"
                )

            # Skip 1 blank row
            current_row += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def strip_section_prefix(name_str: str) -> str:
    """
    Removes a leading pattern like "a. ", "b. ", "n. "
    from the sub-section title.

    Examples:
      "a. Production Product" -> "Production Product"
      "n. Quartz-07" -> "Quartz-07"
    """
    return re.sub(r'^[a-zA-Z]+\.\s*', '', name_str).strip()

def create_placeholder_data(columns):
    """
    Creates a single-row DataFrame with empty string placeholders for each column.
    """
    return pd.DataFrame([{col: "" for col in columns}])

def data_editor_transposed(section_title, variables, selected_date, key_suffix=None):
    """
    Displays a sub-section in transposed form using st.data_editor.
    This is for single-sample data entry (one column per sub-section).
    """
    st.markdown(f"#### {section_title}")

    placeholder_df = create_placeholder_data(variables)
    transposed = placeholder_df.T
    transposed.columns = [str(selected_date)]

    editor_key = f"editor_{section_title}"
    if key_suffix:
        editor_key += f"_{key_suffix}"

    edited = st.data_editor(transposed, num_rows="dynamic", key=editor_key)
    normal_df = edited.T
    return normal_df

def data_editor_multi_samples(
    section_title, variables, selected_date, n_samples=3, sample_labels=None, key_suffix=None
):
    """
    Displays multiple sample columns for a sub-section.
    """
    st.markdown(f"#### {section_title}")

    if not sample_labels or len(sample_labels) != n_samples:
        sample_labels = [f"Sample {i+1}" for i in range(n_samples)]

    df = pd.DataFrame("", index=variables, columns=sample_labels)

    editor_key = f"editor_{section_title}"
    if key_suffix:
        editor_key += f"_{key_suffix}"

    edited_df = st.data_editor(
        df,
        num_rows="fixed",
        use_container_width=True,
        key=editor_key
    )
    return edited_df

def flatten_single_sample_df(df, prefix=None):
    """
    Flatten (1 x N) data to a single row, optionally prefixing column names.
    """
    out_dict = {}
    for col_name in df.columns:
        new_col_name = f"{prefix}_{col_name}" if prefix else col_name
        out_dict[new_col_name] = df.iloc[0][col_name]
    return pd.DataFrame([out_dict])

def flatten_multi_sample_df_agg(df, prefix=None):
    """
    Flatten multi-sample (R x C) by computing avg/min/max for each row (variable).
    """
    out_dict = {}
    for var in df.index:
        vals = []
        for col in df.columns:
            raw_val = df.loc[var, col]
            try:
                numeric_val = float(raw_val)
                vals.append(numeric_val)
            except (ValueError, TypeError):
                pass

        if vals:
            avg_val = np.mean(vals)
            min_val = np.min(vals)
            max_val = np.max(vals)
        else:
            avg_val = min_val = max_val = None

        base_name = f"{prefix}_{var}" if prefix else var
        out_dict[f"{base_name}_avg"] = avg_val
        out_dict[f"{base_name}_min"] = min_val
        out_dict[f"{base_name}_max"] = max_val
    return pd.DataFrame([out_dict])

def layout_sub_sections(sub_sections, selected_date):
    """
    Lay out each sub-section in either single-sample or multi-sample mode,
    returning a dict { sub_section_name -> DataFrame }.
    """
    subsection_data = {}
    cols = None

    for idx, sub_sec in enumerate(sub_sections):
        sub_title = sub_sec["name"]
        variables = sub_sec["variables"]

        wants_multi = sub_sec.get("three_samples", False)
        n_samples = sub_sec.get("n_samples", 1)
        sample_labels = sub_sec.get("sample_labels", None)

        if idx % 2 == 0:
            cols = st.columns(2)

        if wants_multi and n_samples > 1:
            with cols[idx % 2]:
                df_multi = data_editor_multi_samples(
                    section_title=sub_title,
                    variables=variables,
                    selected_date=selected_date,
                    n_samples=n_samples,
                    sample_labels=sample_labels,
                    key_suffix=f"{idx}"
                )
                subsection_data[sub_title] = df_multi
        else:
            with cols[idx % 2]:
                df_normal = data_editor_transposed(
                    section_title=sub_title,
                    variables=variables,
                    selected_date=selected_date,
                    key_suffix=f"{idx}"
                )
                subsection_data[sub_title] = df_normal

    return subsection_data
